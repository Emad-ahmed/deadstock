# views.py
from django.shortcuts import render
import openpyxl
import statistics
import re
from gulfapp.models import ExcelData

def index(request):
    if request.method == "POST":
        excel_files = request.FILES.getlist("excel_files")
        processed_data = process_data(excel_files)

        return render(request, 'import_data.html', {"processed_data": processed_data})

    return render(request, 'import_data.html')

def extract_integer_from_filename(filename):
    # Extracts the first integer found in the filename using regex
    match = re.search(r'\d+', filename)
    return int(match.group()) if match else None

def process_data(excel_files):
    processed_data = {}

    total_files = len(excel_files)

    for excel_file in excel_files:
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb.active

        for row in worksheet.iter_rows(values_only=True):
            product_name = str(row[0])  # Assuming the product column is at index 0
            unit_value = row[1]

            # Check if the unit_value can be converted to a numeric value
            try:
                unit_value = float(unit_value)
            except (ValueError, TypeError):
                # Handle the case where the value is not a valid numeric value
                continue

            file_name = excel_file.name
            file_number = extract_integer_from_filename(file_name)

            if product_name in processed_data:
                processed_data[product_name]["unit_values"].append({"file_number": file_number, "unit_value": unit_value})
            else:
                processed_data[product_name] = {"unit_values": [{"file_number": file_number, "unit_value": unit_value}]}

    # Calculate the total, average, standard deviation, and variance of unit values for each product
    for product_name, data in processed_data.items():
        unit_values = [item["unit_value"] for item in data["unit_values"]]
        total_units = sum(item["unit_value"] for item in data["unit_values"])
        average_unit = sum(item["unit_value"] for item in data["unit_values"]) / total_files if len(data["unit_values"]) > 0 else 0

        # Add 0 values to unit_values to match the length of total_files
        unit_values += [0] * (total_files - len(unit_values))

        # Calculate mean and standard deviation
        mean_unit = sum(unit_values) / len(unit_values)
        sum_squared_diff = sum((x - mean_unit) ** 2 for x in unit_values)
        stdev_unit = (sum_squared_diff / len(unit_values)) ** 0.5
        variance_unit = stdev_unit/average_unit

        # Calculate the percentage of variance
        percentage_variance = (stdev_unit / average_unit) * 100 if average_unit != 0 else 0

        processed_data[product_name]["total_units"] = total_units
        processed_data[product_name]["average_unit"] = average_unit
        processed_data[product_name]["stdev_unit"] = stdev_unit
        processed_data[product_name]["variance_unit"] = variance_unit
        processed_data[product_name]["percentage_variance"] = percentage_variance

    return processed_data



def process_data_mydata(excel_files):
    all_products = set()

    for excel_file in excel_files:
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb.active

        for row in worksheet.iter_rows(values_only=True):
            product_name = str(row[0]).lower()  # Assuming the product column is at index 0
            all_products.add(product_name)
    return all_products

def deadstock(request):
    missing_products = None
    if request.method == "POST":
        excel_files = request.FILES.getlist("excel_files")
        uploaded_products = process_data_mydata(excel_files)
            # Fetch all products from the ExcelData model
        all_excel_data_products = set(ExcelData.objects.values_list('product', flat=True).distinct())

            # Identify products that are in ExcelData model but not in uploaded files
        missing_products = all_excel_data_products - uploaded_products

        return render(request, 'deadstock.html', {"missing_products": missing_products})

   

    return render(request, 'deadstock.html', {"missing_products": missing_products})